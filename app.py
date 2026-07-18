# app.py – Dashboard Streamlit PFE Diabète Type 2
# Exécution : streamlit run app.py
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.stats import chi2_contingency, ttest_ind, fisher_exact
import statsmodels.api as sm
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import glob, os, re, io
import warnings
warnings.filterwarnings('ignore')

# ======================================================================
# 1. CONFIGURATION
# ======================================================================
st.set_page_config(
    page_title="PFE Diabète – Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)
st.title("📊 Pipeline Analytique & Prédictif – Diabète Type 2")
st.subheader("Institut Pasteur du Maroc – Casablanca & Jadida")
st.markdown("**Statistiques (OR, ACP, Heatmap, Stratification, SHAP) + ML**")
st.divider()

# ======================================================================
# 2. HELPER : FIGURE → BYTES PNG
# ======================================================================
def fig_vers_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    return buf.getvalue()

# ======================================================================
# 3. FONCTIONS STATISTIQUES
# ======================================================================
def format_p(p):
    if p is None or pd.isna(p): return 'N/A'
    return '≤ 0,001' if p < 0.001 else f'{p:.3f}'

def test_continu(data, var):
    g1 = data[data['cible']==1][var].dropna()
    g0 = data[data['cible']==0][var].dropna()
    if len(g1)<2 or len(g0)<2: return None,None,None,None,None
    _,p = ttest_ind(g1, g0, equal_var=False)
    return g1.mean(), g1.std(), g0.mean(), g0.std(), p

def test_cat(data, var):
    sous = data[[var,'cible']].dropna()
    tab  = pd.crosstab(sous[var], sous['cible'])
    if tab.shape[0]<2 or tab.shape[1]<2: return None,None
    _,p,_,_ = chi2_contingency(tab)
    return tab, p

def calc_or_brut(data, col):
    tab = pd.crosstab(data[col], data['cible'])
    if tab.shape!=(2,2): return None,None,None,None
    try:
        or_,p = fisher_exact(tab)
        a,b,c,d = tab.iloc[0,0],tab.iloc[0,1],tab.iloc[1,0],tab.iloc[1,1]
        if 0 in [a,b,c,d]: return None,None,None,None
        se = np.sqrt(1/a+1/b+1/c+1/d)
        return or_, np.exp(np.log(or_)-1.96*se), np.exp(np.log(or_)+1.96*se), p
    except: return None,None,None,None

def reg_log(data, col):
    sub = data[[col,'Age','sexe','cible']].dropna()
    if len(sub)<10 or sub[col].nunique()<2 or sub['cible'].nunique()<2:
        return None,None,None,None
    sub = sub.copy()
    sub['sexe_b'] = sub['sexe'].map({'Homme':1,'Femme':0})
    X = sm.add_constant(sub[['Age','sexe_b',col]])
    y = sub['cible']
    try:
        m = sm.Logit(y,X).fit(disp=0)
        c  = m.params[col]; ci = m.conf_int().loc[col]; p = m.pvalues[col]
        return np.exp(c), np.exp(ci[0]), np.exp(ci[1]), p
    except: return None,None,None,None

def style_ws(ws, df_t, titre):
    H  = PatternFill(start_color='1F5C99',end_color='1F5C99',fill_type='solid')
    R1 = PatternFill(start_color='EAF2FB',end_color='EAF2FB',fill_type='solid')
    R2 = PatternFill(start_color='FFFFFF',end_color='FFFFFF',fill_type='solid')
    bd = Side(style='thin',color='CCCCCC')
    br = Border(left=bd,right=bd,top=bd,bottom=bd)
    nc = len(df_t.columns)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=nc)
    ws['A1']=titre
    ws['A1'].font=Font(name='Arial',bold=True,size=11,color='1F5C99')
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[1].height=30
    for j,col in enumerate(df_t.columns,1):
        c=ws.cell(row=2,column=j,value=col)
        c.fill=H; c.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=br
    ws.row_dimensions[2].height=35
    for i,(_,row) in enumerate(df_t.iterrows(),3):
        fill=R1 if i%2==0 else R2
        for j,val in enumerate(row,1):
            c=ws.cell(row=i,column=j,value=str(val) if not pd.isna(val) else '')
            c.fill=fill; c.font=Font(name='Arial',size=10)
            c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            c.border=br
        ws.row_dimensions[i].height=22
    for j,col in enumerate(df_t.columns,1):
        ml=max(len(str(col)),df_t[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(j)].width=min(ml+4,40)

# ======================================================================
# 4. PIPELINE COMPLET (mode auto) – VERSION CORRIGÉE
# ======================================================================
NOMS = {
    'Viande rouge_bin':'Viande rouge','Poulet_bin':'Poulet','poisson_bin':'Poisson',
    'Hot dog_bin':'Hot dog','Saucisson_bin':'Saucisson','Légumes_bin':'Légumes',
    'Fruits_bin':'Fruits','Œufs_bin':'Œufs','Lait_bin':'Lait',
    'Produits laitiers_bin':'Produits laitiers','Café_bin':'Café','Thé_bin':'Thé',
    'Pattes_bin':'Pâtes / Féculents','Sucreries_bin':'Sucreries',
    'Boissons énergétiques_bin':'Boissons énergétiques',
    'Produits  riche en matière grasse_bin':'Produits riches en MG',
    'Soupe_bin':'Soupe',"Huile d'olive_bin":"Huile d'olive",
    'Fruits secs_bin':'Fruits secs','Céréales_bin':'Céréales',
    "Nourriture à l'extérieur_bin":'Restauration hors domicile',
    'Fast Food_bin':'Fast Food','Activité physique_bin':'Activité physique',
    'smoking statuts_bin':'Tabagisme',
}
EXCLUS_BIN = ['Nombre de repas_bin','Eau_bin']

@st.cache_data(show_spinner=False)
def run_pipeline(file_bytes):
    """
    Pipeline complet depuis le fichier Excel brut (4 feuilles JM/CM/JT/CT).
    Retourne : df, tableaux (dict), figures_bytes (dict de PNG bytes), rapport, excel_bytes.
    """
    # ── 1. Chargement des 4 feuilles avec conversion sélective ──────
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    frames = []
    # Liste des colonnes numériques à convertir en float
    numeric_cols = ['Age', 'Poids', 'Hauteur', 'Systolique', 'diastolique', 
                    'Hb1AC', 'Glycémie', 'Urée', 'Acide urique', 'Créatinine',
                    '(GPT)ALAT', '(GOT)ASAT', 'CT', 'TG', 'HDL', 'LDL']
    
    for nom, cible, ville in [('JM',1,'Jadida'),('CM',1,'Casablanca'),
                               ('JT',0,'Jadida'),('CT',0,'Casablanca')]:
        try:
            df_ = xl.parse(nom, header=1)
            # Supprimer les lignes/colonnes totalement vides
            df_ = df_.dropna(how='all').dropna(axis=1, how='all')
            
            # Nettoyage des colonnes
            for col in df_.columns:
                # Pour les colonnes numériques : forcer en float
                if col in numeric_cols:
                    df_[col] = pd.to_numeric(df_[col], errors='coerce')
                # Pour les autres : convertir en chaîne et remplacer les valeurs vides par NaN
                else:
                    df_[col] = df_[col].astype(str).replace('nan', np.nan).replace('', np.nan)
            
            df_['cible'] = cible
            df_['Ville'] = ville
            frames.append(df_)
        except Exception:
            pass
    if not frames:
        raise ValueError("Aucune feuille valide (JM/CM/JT/CT).")
    df = pd.concat(frames, ignore_index=True)
    df['cible'] = df['cible'].astype(int)

    # ── 2. Nettoyages ─────────────────────────────────────────────────
    df['sexe_clean'] = df['sexe'].astype(str).str.strip().str.title()
    df['sexe_clean'] = df['sexe_clean'].where(df['sexe_clean'].isin(['Femme','Homme']),np.nan)
    df['ville_clean'] = df['Ville'].astype(str).str.strip().str.title()
    df['milieu_clean'] = df['Milieu de vie'].astype(str).str.strip().str.lower().map(
        {'urbain':'Urbain','rural':'Rural','rurale':'Rural'})

    for col_r,col_n in [
        ('Hérédité','heredite_clean'),('Hypertension','hta_clean'),
        ('Dyslipidémie','dyslipid_clean'),('consanguinité','consang_clean'),
        ('Cardiovasculaire','cv_clean'),('Nephropathie/rénale','nephro_clean'),
        ('Rétinopathie','retino_clean'),('Neuropathie','neuro_clean'),
        ('Pied diabétique','pied_clean'),
    ]:
        if col_r in df.columns:
            df[col_n] = df[col_r].astype(str).str.strip().str.lower().map(
                {'oui':'Oui','oui ':'Oui','non':'Non','non ':'Non'})

    def _psycho(v):
        if pd.isna(v): return np.nan
        v=str(v).strip().lower()
        if v in ['non','non ']: return 'Non'
        if any(m in v for m in ['stress','anxiété','oui','humeur']): return 'Oui'
        return np.nan
    if 'Psychologique' in df.columns:
        df['psycho_clean'] = df['Psychologique'].apply(_psycho)

    # IMC
    poids = pd.to_numeric(df.get('Poids'),errors='coerce')
    haut  = pd.to_numeric(df.get('Hauteur'),errors='coerce').apply(
        lambda x: x/100 if pd.notna(x) and x>3 else x)
    df['IMC_calc'] = (poids/(haut**2)).where(lambda x:(x>=12)&(x<=70))

    # TA
    def _ta(v,seuil,lo,hi):
        val=pd.to_numeric(v,errors='coerce')
        if pd.isna(val): return np.nan
        if val<seuil: val*=10
        return val if lo<=val<=hi else np.nan
    if 'Systolique'  in df.columns: df['sys_mmhg']  = df['Systolique'].apply(lambda v:_ta(v,25,70,250))
    if 'diastolique' in df.columns: df['dias_mmhg'] = df['diastolique'].apply(lambda v:_ta(v,15,40,130))

    # Binarisation variables alimentaires
    FREQ_OUI=['Souvent','Souvent ','souvent','De temps en temps']
    alim=['Viande rouge','Poulet','poisson','Hot dog','Saucisson','Légumes','Fruits',
          'Œufs','Lait','Produits laitiers','Café','Thé','Pattes','Sucreries',
          'Boissons énergétiques','Produits  riche en matière grasse','Soupe',
          "Huile d'olive",'Fruits secs','Céréales',"Nourriture à l'extérieur",'Fast Food']
    for col in alim:
        bn=col+'_bin'
        if bn not in df.columns and col in df.columns:
            df[bn]=df[col].apply(lambda x:1 if str(x).strip() in FREQ_OUI else (0 if pd.notna(x) else np.nan))
    if 'Activité physique' in df.columns and 'Activité physique_bin' not in df.columns:
        df['Activité physique_bin']=df['Activité physique'].apply(
            lambda x:1 if str(x).strip().lower() in ['oui','actif'] else (0 if pd.notna(x) else np.nan))
    if 'smoking statuts' in df.columns and 'smoking statuts_bin' not in df.columns:
        df['smoking statuts_bin']=df['smoking statuts'].apply(
            lambda x:1 if str(x).strip().lower() in ['fumeur','oui'] else (0 if pd.notna(x) else np.nan))

    bin_cols=[c for c in df.columns if c.endswith('_bin') and c not in EXCLUS_BIN]
    N=len(df); N_MAL=(df['cible']==1).sum(); N_TEM=(df['cible']==0).sum()

    # ── 3. Tableaux ───────────────────────────────────────────────────
    tableaux={}

    # T1 – Caractéristiques générales
    rows=[]
    for var,lbl,cont in [('Age','Âge (moy ± ET)',True),('IMC_calc','IMC (moy ± ET)',True)]:
        if var not in df.columns: continue
        r=test_continu(df,var)
        if r[0]: rows.append([lbl,f'{r[0]:.1f}±{r[1]:.1f}',f'{r[2]:.1f}±{r[3]:.1f}',format_p(r[4]),'t-test'])
    for v,l in [('sexe_clean','Sexe'),('Ville','Ville'),('milieu_clean','Milieu')]:
        if v not in df.columns: continue
        tab,p=test_cat(df,v)
        if tab is None: continue
        for cat in tab.index:
            pm=100*tab.loc[cat,1]/tab[1].sum() if 1 in tab.columns else 0
            pt=100*tab.loc[cat,0]/tab[0].sum() if 0 in tab.columns else 0
            rows.append([f'{l}:{cat}',f'{pm:.1f}%',f'{pt:.1f}%',format_p(p),'Chi²'])
    for v,l in [('heredite_clean','Hérédité (% Oui)'),('hta_clean','Hypertension (% Oui)'),
                ('dyslipid_clean','Dyslipidémie (% Oui)'),('psycho_clean','Stress/Anxiété (% Oui)')]:
        if v not in df.columns: continue
        tab,p=test_cat(df,v)
        if tab is None or 'Oui' not in tab.index: continue
        pm=100*tab.loc['Oui',1]/tab[1].sum() if 1 in tab.columns else 0
        pt=100*tab.loc['Oui',0]/tab[0].sum() if 0 in tab.columns else 0
        rows.append([l,f'{pm:.1f}%',f'{pt:.1f}%',format_p(p),'Chi²'])
    tableaux['T1']=pd.DataFrame(rows,columns=['Paramètre',f'Malades(n={N_MAL})',f'Témoins(n={N_TEM})','p','Test'])

    # T2 – Prévalence
    rows2=[]
    for col in bin_cols:
        nom=NOMS.get(col,col.replace('_bin',''))
        pm=df[df['cible']==1][col].mean()*100; pt=df[df['cible']==0][col].mean()*100
        _,p=test_cat(df,col)
        rows2.append([nom,f'{pm:.1f}',f'{pt:.1f}',format_p(p) if p else 'N/A'])
    tableaux['T2']=pd.DataFrame(rows2,columns=['Facteur','Malades(%)','Témoins(%)','p'])

    # T3 – OR bruts
    rows3=[]
    for col in bin_cols:
        nom=NOMS.get(col,col.replace('_bin',''))
        or_,cil,cih,p=calc_or_brut(df,col)
        if or_: rows3.append([nom,f'{or_:.2f}({cil:.2f}-{cih:.2f})',format_p(p),p])
    t3f=pd.DataFrame(rows3,columns=['Facteur','OR brut(IC95%)','p','_p']).sort_values('_p')
    tableaux['T3']=t3f[['Facteur','OR brut(IC95%)','p']].reset_index(drop=True)

    # T4 – OR ajustés
    rows4=[]
    for col in bin_cols:
        nom=NOMS.get(col,col.replace('_bin',''))
        or_a,cil_a,cih_a,p_a=reg_log(df,col)
        if or_a: rows4.append([nom,or_a,cil_a,cih_a,f'{or_a:.2f}({cil_a:.2f}-{cih_a:.2f})',format_p(p_a),p_a])
    t4f=pd.DataFrame(rows4,columns=['Facteur','_or','_ci_low','_ci_high','OR ajusté(IC95%)','p','_p']).sort_values('_p')
    tableaux['T4']=t4f[['Facteur','OR ajusté(IC95%)','p']].reset_index(drop=True)

    # T5 – Médical malades
    df_m=df[df['cible']==1].copy()
    def _ext(v):
        if pd.isna(v): return np.nan
        v=str(v).lower(); m=re.search(r'(\d+\.?\d*)',v)
        if not m: return np.nan
        val=float(m.group(1)); return round(val/12,2) if 'mois' in v else val
    if 'Ancienneté du DT2' in df_m.columns:
        df_m['dur']=df_m['Ancienneté du DT2'].apply(_ext)
        nd=df_m['dur'].notna().sum()
    else:
        df_m['dur']=np.nan; nd=0
    rows5=[[f'Durée DT2 (moy±ET)',f"{df_m['dur'].mean():.1f}±{df_m['dur'].std():.1f}" if nd>0 else 'N/A',f'n={nd}']]
    if 'Hb1AC' in df_m.columns:
        df_m['hba']=pd.to_numeric(df_m['Hb1AC'],errors='coerce')
        n_hba=df_m['hba'].notna().sum()
        rows5.append(['HbA1c(%) moy±ET',f"{df_m['hba'].mean():.1f}±{df_m['hba'].std():.1f}",f'n={n_hba}'])
    for col,lbl in [('hta_clean','Hypertension(% Oui)'),('dyslipid_clean','Dyslipidémie(% Oui)'),('psycho_clean','Stress/Anxiété(% Oui)')]:
        if col not in df_m.columns: continue
        n_oui=(df_m[col]=='Oui').sum(); n_tot=df_m[col].notna().sum()
        if n_tot>0: rows5.append([lbl,f'{n_oui}({100*n_oui/n_tot:.1f}%)',f'n={n_tot}'])
    tableaux['T5']=pd.DataFrame(rows5,columns=['Paramètre',f'Malades(n={len(df_m)})','Note'])

    # T6 – Complications
    COMP=[('cv_clean','Cardiovasculaire'),('retino_clean','Rétinopathie'),
          ('nephro_clean','Néphropathie'),('neuro_clean','Neuropathie'),
          ('pied_clean','Pied diabétique'),('psycho_clean','Psychologique')]
    rows6=[]
    for var,lbl in COMP:
        if var not in df.columns: continue
        nm=(df_m[var]=='Oui').sum() if var in df_m.columns else 0
        tm=df_m[var].notna().sum() if var in df_m.columns else 1
        nt=(df[df['cible']==0][var]=='Oui').sum(); tt=df[df['cible']==0][var].notna().sum()
        _,p=test_cat(df,var)
        rows6.append([lbl,f'{nm}({100*nm/tm:.1f}%)',f'{nt}({100*nt/tt:.1f}%)',format_p(p) if p else 'N/A'])
    tableaux['T6']=pd.DataFrame(rows6,columns=['Complication',f'Malades(n={len(df_m)})',f'Témoins(n={N_TEM})','p'])

    # T7 – Stratification par ville (même si aucun facteur significatif)
    facteurs_sig = t4f[t4f['_p'] < 0.05]['Facteur'].tolist()
    if facteurs_sig:
        cols_sig = [c for c in bin_cols if NOMS.get(c, c.replace('_bin','')) in facteurs_sig]
        rows7 = []
        for col in cols_sig:
            nom = NOMS.get(col, col.replace('_bin',''))
            ligne = [str(nom)]
            for ville in ['Casablanca','Jadida']:
                df_v = df[df['Ville'] == ville]
                or_v, cil_v, cih_v, p_v = reg_log(df_v, col)
                if or_v is not None:
                    ligne.append(f'{or_v:.2f}({cil_v:.2f}-{cih_v:.2f})')
                    ligne.append(format_p(p_v))
                else:
                    ligne.append('—')
                    ligne.append('N/A')
            rows7.append(ligne)
        tableaux['T7'] = pd.DataFrame(rows7, columns=['Facteur', 'OR Casa(IC95%)', 'p Casa', 'OR Jadida(IC95%)', 'p Jadida'])
    else:
        # Si aucun facteur significatif, on crée un DataFrame avec un message
        tableaux['T7'] = pd.DataFrame([['Aucun facteur significatif (p<0,05) – stratification non réalisée.', '', '', '', '']],
                                      columns=['Facteur', 'OR Casa(IC95%)', 'p Casa', 'OR Jadida(IC95%)', 'p Jadida'])

    # ── 4. FIGURES → BYTES PNG ──────────────────────────────────────
    figs_bytes={}

    # Forest plot
    signif = t4f[t4f['_p'] < 0.05]
    if not signif.empty:
        plot_df = signif.sort_values('_or').reset_index(drop=True)
        fig, ax = plt.subplots(figsize=(9, max(4, 0.48 * len(plot_df))))
        y_pos = list(range(len(plot_df)))
        ax.errorbar(plot_df['_or'], y_pos,
                    xerr=[plot_df['_or'] - plot_df['_ci_low'], plot_df['_ci_high'] - plot_df['_or']],
                    fmt='none', capsize=5, color='#1F5C99', elinewidth=1.5)
        for i, row in plot_df.iterrows():
            ax.scatter(row['_or'], i, color='#E53935' if row['_or'] > 1 else '#1565C0', zorder=5, s=70)
        ax.axvline(1, linestyle='--', color='red', alpha=0.6)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(plot_df['Facteur'], fontsize=10)
        ax.set_xlabel('OR ajusté (IC95%) – log', fontsize=11)
        ax.set_xscale('log')
        ax.set_title('Facteurs environnementaux associés au DT2\n(ajustés âge+sexe, p<0,05)',
                     fontsize=12, fontweight='bold', color='#1F5C99')
        ax.grid(axis='x', alpha=0.3, linestyle=':')
        plt.tight_layout()
        figs_bytes['forest'] = fig_vers_bytes(fig)
    else:
        # Créer une figure avec un message explicite
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.text(0.5, 0.5, 'Aucun facteur avec p < 0,05\nForest plot non généré.',
                ha='center', va='center', fontsize=14, color='#1F5C99', fontweight='bold')
        ax.axis('off')
        plt.tight_layout()
        figs_bytes['forest'] = fig_vers_bytes(fig)

    # ACP (scree + individus + cercle) → UNE SEULE figure, 3 panneaux
    X_acp=df[bin_cols].fillna(0)
    X_sc=StandardScaler().fit_transform(X_acp)
    pca_full=PCA().fit(X_sc)
    pca=PCA(n_components=2); sc2=pca.fit_transform(X_sc)
    df['PC1']=sc2[:,0]; df['PC2']=sc2[:,1]
    v1,v2=pca.explained_variance_ratio_
    fig,axes=plt.subplots(1,3,figsize=(20,6))
    var_exp=pca_full.explained_variance_ratio_[:10]*100
    axes[0].bar(range(1,11),var_exp,color='#1F5C99',alpha=0.8,edgecolor='white')
    axes[0].plot(range(1,11),np.cumsum(var_exp),'r--o',markersize=5,label='Cumulé')
    axes[0].set_title('Scree plot',fontweight='bold'); axes[0].legend()
    for cv,lbl,col in [(0,'Témoins','#1565C0'),(1,'Malades','#E53935')]:
        sub=df[df['cible']==cv]
        axes[1].scatter(sub['PC1'],sub['PC2'],c=col,label=lbl,alpha=0.4,s=15)
    axes[1].set_xlabel(f'PC1({v1:.1%})'); axes[1].set_ylabel(f'PC2({v2:.1%})')
    axes[1].set_title('Projection des individus',fontweight='bold'); axes[1].legend()
    axes[1].grid(alpha=0.3)
    ld=pca.components_.T; nv=[NOMS.get(c,c.replace('_bin','')) for c in bin_cols]
    axes[2].add_artist(plt.Circle((0,0),1,fill=False,linestyle='--',color='gray',alpha=0.5))
    for i,n in enumerate(nv):
        axes[2].arrow(0,0,ld[i,0]*0.88,ld[i,1]*0.88,head_width=0.04,head_length=0.04,
                      fc='#1F5C99',ec='#1F5C99',alpha=0.7)
        axes[2].text(ld[i,0]*1.07,ld[i,1]*1.07,n,fontsize=7.5,ha='center',va='center')
    axes[2].set_xlim(-1.3,1.3); axes[2].set_ylim(-1.3,1.3); axes[2].set_aspect('equal')
    axes[2].set_title('Cercle des corrélations',fontweight='bold'); axes[2].grid(alpha=0.2)
    plt.tight_layout()
    figs_bytes['acp'] = fig_vers_bytes(fig)

    # Heatmap
    corr=df[bin_cols].corr()
    fig,ax=plt.subplots(figsize=(14,12))
    im=ax.imshow(corr.values,cmap='RdYlBu_r',vmin=-1,vmax=1,aspect='auto')
    noms_c=[NOMS.get(c,c.replace('_bin','')) for c in bin_cols]
    ax.set_xticks(range(len(bin_cols))); ax.set_yticks(range(len(bin_cols)))
    ax.set_xticklabels(noms_c,rotation=45,ha='right',fontsize=8)
    ax.set_yticklabels(noms_c,fontsize=8)
    for i in range(len(bin_cols)):
        for j in range(len(bin_cols)):
            val=corr.values[i,j]
            if i!=j and abs(val)>=0.3:
                ax.text(j,i,f'{val:.2f}',ha='center',va='center',fontsize=6,
                        color='white' if abs(val)>0.6 else 'black',fontweight='bold')
    plt.colorbar(im,ax=ax,shrink=0.8,label='Corrélation')
    ax.set_title('Matrice de corrélation – Facteurs environnementaux',
                 fontsize=12,fontweight='bold',color='#1F5C99',pad=15)
    plt.tight_layout()
    figs_bytes['heatmap'] = fig_vers_bytes(fig)

    # SHAP
    try:
        import shap
        df['sexe_num']=df['sexe'].map({'Homme':1,'Femme':0})
        feat=bin_cols+['Age','sexe_num']
        noms_sh=[NOMS.get(c,c.replace('_bin','')) for c in bin_cols]+['Âge','Sexe']
        sub_s=df[feat+['cible']].dropna()
        X_s=sub_s[feat].values; y_s=sub_s['cible'].values
        rf=RandomForestClassifier(n_estimators=200,max_depth=6,random_state=42,
                                  class_weight='balanced',n_jobs=-1)
        rf.fit(X_s,y_s)
        exp=shap.TreeExplainer(rf)
        sv=exp.shap_values(X_s)
        sv_m = sv[:,:,1] if (hasattr(sv,'ndim') and sv.ndim==3) else (sv[1] if isinstance(sv,list) else sv)
        mabs=np.abs(sv_m).mean(axis=0); idx_s=np.argsort(mabs)[::-1]
        top=min(20,len(feat)); idx_top=idx_s[:top]
        fig,ax=plt.subplots(figsize=(10,7))
        ax.barh([noms_sh[i] for i in idx_top[::-1]],[mabs[i] for i in idx_top[::-1]],
                color='#1565C0',edgecolor='white',alpha=0.85)
        ax.set_xlabel('Importance SHAP moyenne (|valeur SHAP|)',fontsize=11)
        ax.set_title('Interprétabilité SHAP – Importance des facteurs\n(Random Forest)',
                     fontsize=12,fontweight='bold',color='#1F5C99')
        ax.grid(axis='x',alpha=0.3,linestyle=':')
        plt.tight_layout()
        figs_bytes['shap'] = fig_vers_bytes(fig)
    except Exception:
        pass

    # ── 5. Export Excel en mémoire ────────────────────────────────────
    wb_out=Workbook()
    def _add_sheet(wb_, key, title, premier=False):
        ws_=wb_.active if premier else wb_.create_sheet(title)
        if not premier: ws_.title=title
        else: ws_.title=title
        style_ws(ws_,tableaux[key],f'Tableau – {title}')

    _add_sheet(wb_out,'T1','T1_Caracteristiques',True)
    for k,t in [('T2','T2_Prevalence'),('T3','T3_Univarie'),('T4','T4_Multivarie'),
                ('T5','T5_Medical'),('T6','T6_Complications'),('T7','T7_Stratification')]:
        _add_sheet(wb_out,k,t)
    buf=io.BytesIO(); wb_out.save(buf)
    excel_bytes=buf.getvalue()

    # ── 6. Mini-rapport ──────────────────────────────────────────────
    n_sig4=(t4f['_p']<0.05).sum()
    rapport=(f"RAPPORT PRÉLIMINAIRE – PIPELINE PFE DIABÈTE TYPE 2\n"
             f"Institut Pasteur du Maroc | n={N} patients\n{'='*60}\n\n"
             f"1. POPULATION\n"
             f"   Malades : {N_MAL} ({100*N_MAL/N:.1f}%) | Témoins : {N_TEM} ({100*N_TEM/N:.1f}%)\n"
             f"   Âge moy Malades : {df[df['cible']==1]['Age'].mean():.1f} ans\n\n"
             f"2. FACTEURS SIGNIFICATIFS EN MULTIVARIÉ\n"
             f"   {n_sig4}/{len(tableaux['T4'])} facteurs (p<0,05)\n")

    return df, tableaux, figs_bytes, rapport, excel_bytes

# ======================================================================
# 5. CHARGEMENT MODE RAPIDE – CORRECTION : header=1 pour utiliser la 2e ligne comme en-tête
# ======================================================================
@st.cache_data
def load_data_rapide():
    for fname in ["data/diabete_dashboard_ready.csv","data/diabete_clean_general.csv"]:
        if os.path.exists(fname):
            df=pd.read_csv(fname); df['cible']=df['cible'].astype(int)
            return df, fname
    st.error("❌ Aucun fichier CSV trouvé."); st.stop()

@st.cache_data
def load_sheet(sh):
    try:
        # On utilise header=1 pour prendre la 2e ligne comme noms de colonnes
        df = pd.read_excel("data/Tableaux_complets_PFE_v2.xlsx", sheet_name=sh, header=1)
        return df
    except Exception:
        return None

@st.cache_data
def load_ml():
    try: return pd.read_csv("data/resultats_modeles_ML.csv")
    except: return None

# ======================================================================
# 6. BARRE LATÉRALE – MODE + FILTRES
# ======================================================================
st.sidebar.header("⚙️ Configuration")
mode=st.sidebar.radio("📌 Mode",
    ["⚡ Mode rapide (fichiers pré-générés)","🔄 Mode auto (upload Excel brut)"])

for k in ['df','tableaux','figs_bytes','rapport','excel_bytes','mode_auto','source','upl_name']:
    if k not in st.session_state: st.session_state[k]=None if k!='mode_auto' else False

if mode=="⚡ Mode rapide (fichiers pré-générés)":
    st.session_state['mode_auto']=False
    if st.session_state['df'] is None or st.session_state['mode_auto']:
        df_r,src=load_data_rapide()
        st.session_state['df']=df_r; st.session_state['source']=src
    st.sidebar.success("✅ Mode rapide")
else:
    st.session_state['mode_auto']=True
    uploaded=st.sidebar.file_uploader("📂 Excel brut (JM/CM/JT/CT)",type=["xlsx"])
    if uploaded:
        if st.session_state.get('upl_name')!=uploaded.name:
            with st.spinner("⏳ Pipeline en cours (1-2 min)..."):
                try:
                    df_r,tbx,fbs,rpt,excl=run_pipeline(uploaded.read())
                    st.session_state.update({'df':df_r,'tableaux':tbx,'figs_bytes':fbs,
                        'rapport':rpt,'excel_bytes':excl,'upl_name':uploaded.name,
                        'source':f"Excel : {uploaded.name}"})
                    st.success("✅ Pipeline terminé !")
                except Exception as e:
                    st.error(f"❌ Erreur : {e}"); st.stop()
    else:
        st.info("📤 Uploadez le fichier Excel brut."); st.stop()

if st.session_state['mode_auto'] and st.session_state.get('excel_bytes'):
    st.sidebar.download_button("📥 Télécharger Excel",
        data=st.session_state['excel_bytes'],
        file_name="Tableaux_complets_PFE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

df=st.session_state['df']
if df is None: st.warning("⚠️ Aucune donnée."); st.stop()
st.sidebar.caption(f"📄 {st.session_state.get('source','')}")

# Filtres
st.sidebar.header("🔍 Filtres")
villes=sorted(df['Ville'].dropna().unique().tolist())
vf=st.sidebar.multiselect("🌍 Ville",options=villes,default=villes)
amin,amax=int(df['Age'].min()),int(df['Age'].max())
ar=st.sidebar.slider("📅 Âge",amin,amax,(amin,amax))
sf=st.sidebar.selectbox("⚧️ Sexe",["Tous","Femme","Homme"])

if not vf:
    st.sidebar.warning("⚠️ Sélectionnez une ville.")
    mask=pd.Series(False,index=df.index)
else:
    mask=df['Ville'].isin(vf)&df['Age'].between(ar[0],ar[1])
    if sf!="Tous": mask=mask&(df['sexe']==sf)

df_f=df[mask]
st.sidebar.divider()
st.sidebar.metric("👥 Total",len(df_f))
st.sidebar.metric("🟥 Malades",(df_f['cible']==1).sum())
st.sidebar.metric("🟦 Témoins",(df_f['cible']==0).sum())
VIDE=len(df_f)==0
if VIDE: st.warning("⚠️ Aucun patient avec ces filtres.")

# ======================================================================
# 7. ONGLETS
# ======================================================================
MA=st.session_state['mode_auto']

tab1,tab2,tab3,tab4,tab5,tab6,tab7=st.tabs([
    "📈 Exploratoire","📋 Tableaux","🌳 Forest Plot",
    "📉 ACP","🤖 Machine Learning","📄 Mini-rapport","🔬 Analyses approfondies"
])

# ── Onglet 1 : EDA ────────────────────────────────────────────────────
with tab1:
    st.markdown("### Analyse descriptive interactive")
    if VIDE:
        st.info("Aucune donnée.")
    else:
        c1,c2=st.columns(2)
        with c1:
            fig,ax=plt.subplots(figsize=(6,4))
            nm=(df_f['cible']==1).sum(); nt=(df_f['cible']==0).sum()
            ax.pie([nm,nt],
                   labels=[f"Malades\n({nm}-{100*nm/len(df_f):.1f}%)",
                            f"Témoins\n({nt}-{100*nt/len(df_f):.1f}%)"],
                   colors=['#1F5C99','#E87722'],autopct='%1.1f%%')
            ax.set_title("Répartition par statut")
            st.pyplot(fig); plt.close(fig)

            fig,ax=plt.subplots(figsize=(6,4))
            data_age=[df_f[df_f['cible']==1]['Age'].dropna(),
                      df_f[df_f['cible']==0]['Age'].dropna()]
            try:
                ax.boxplot(data_age,tick_labels=['Malades','Témoins'],
                           patch_artist=True,boxprops=dict(facecolor='#EAF2FB'),
                           medianprops=dict(color='#1F5C99',lw=2))
            except TypeError:
                ax.boxplot(data_age,labels=['Malades','Témoins'],
                           patch_artist=True,boxprops=dict(facecolor='#EAF2FB'),
                           medianprops=dict(color='#1F5C99',lw=2))
            ax.set_ylabel("Âge (années)"); ax.set_title("Âge par statut")
            ax.grid(axis='y',alpha=0.3); st.pyplot(fig); plt.close(fig)

        with c2:
            col_sx='sexe_clean' if 'sexe_clean' in df_f.columns else 'sexe'
            fig,ax=plt.subplots(figsize=(6,4))
            ct=pd.crosstab(df_f[col_sx],df_f['cible']).reindex(columns=[0,1],fill_value=0)
            ct.columns=['Témoins','Malades']
            ct.plot(kind='bar',ax=ax,color=['#1F5C99','#E87722'],edgecolor='white')
            ax.set_title("Sexe par statut"); ax.tick_params(axis='x',rotation=0)
            st.pyplot(fig); plt.close(fig)

            fig,ax=plt.subplots(figsize=(6,4))
            ct=pd.crosstab(df_f['Ville'],df_f['cible']).reindex(columns=[0,1],fill_value=0)
            ct.columns=['Témoins','Malades']
            ct.plot(kind='bar',ax=ax,color=['#1F5C99','#E87722'],edgecolor='white')
            ax.set_title("Ville par statut"); ax.tick_params(axis='x',rotation=0)
            st.pyplot(fig); plt.close(fig)

        st.divider()
        cdsp=['Age']+(['IMC_calc'] if 'IMC_calc' in df_f.columns else [])
        st.markdown("#### Statistiques descriptives")
        st.dataframe(df_f[cdsp].describe().T.style.format("{:.1f}"),width='stretch')

# ── Onglet 2 : Tableaux ───────────────────────────────────────────────
with tab2:
    st.markdown("### Tableaux statistiques (7 feuilles)")
    if MA:
        tbx=st.session_state.get('tableaux',{})
        for key,title in [('T1',"📋 T1 – Caractéristiques"),('T2',"📊 T2 – Prévalence"),
                          ('T3',"📈 T3 – OR bruts"),('T4',"📈 T4 – OR ajustés"),
                          ('T5',"🩺 T5 – Médical"),('T6',"⚠️ T6 – Complications"),
                          ('T7',"🗺️ T7 – Stratification")]:
            dft=tbx.get(key)
            if dft is not None:
                with st.expander(title, expanded=False):
                    if dft.empty:
                        st.info("Ce tableau ne contient pas de données (aucun facteur significatif ou cas particulier).")
                    else:
                        st.dataframe(dft, width='stretch')
            else:
                st.warning(f"Tableau {key} non disponible.")
    else:
        absent=True
        for sh,title in [
            ("T1_Caracteristiques","📋 T1"),("T2_Prevalence_env","📊 T2"),
            ("T3_Univarie_OR_bruts","📈 T3"),("T4_Multivarie_OR_ajustes","📈 T4"),
            ("T5_Medical_malades","🩺 T5"),("T6_Complications","⚠️ T6"),
            ("T7_Stratification_ville","🗺️ T7")]:
            dft=load_sheet(sh)
            if dft is not None:
                absent=False
                with st.expander(title,expanded=False): st.dataframe(dft,width='stretch')
        if absent: st.warning("⚠️ Tableaux_complets_PFE_v2.xlsx introuvable.")

# ── Onglet 3 : Forest Plot ────────────────────────────────────────────
with tab3:
    st.markdown("### Forest plot – OR ajustés (p<0,05)")
    if MA:
        fb=st.session_state.get('figs_bytes',{}).get('forest')
        if fb:
            st.image(fb, use_container_width=True)
        else:
            st.warning("⚠️ Forest plot non disponible.")
    else:
        if os.path.exists("images/forest_plot_OR_ajustes.png"):
            st.image(Image.open("images/forest_plot_OR_ajustes.png"), use_container_width=True)
        else:
            st.warning("⚠️ forest_plot_OR_ajustes.png introuvable.")
    dft4=st.session_state.get('tableaux',{}).get('T4') if MA else load_sheet("T4_Multivarie_OR_ajustes")
    if dft4 is not None:
        st.markdown("#### Détail OR ajustés")
        st.dataframe(dft4,width='stretch')

# ── Onglet 4 : ACP ───────────────────────────────────────────────────
with tab4:
    st.markdown("### Analyse en Composantes Principales (ACP)")
    st.caption("Note : ACP exploratoire sur variables binaires. ACM plus rigoureuse en théorie.")
    if MA:
        fb=st.session_state.get('figs_bytes',{}).get('acp')
        if fb: st.image(fb,use_container_width=True)
        else:  st.warning("⚠️ ACP non disponible.")
    else:
        if os.path.exists("images/ACP_complet.png"):
            st.image(Image.open("images/ACP_complet.png"),use_container_width=True)
        else: st.warning("⚠️ ACP_complet.png introuvable.")

# ── Onglet 5 : Machine Learning ──────────────────────────────────────
with tab5:
    st.markdown("### 🤖 Modélisation prédictive (notebook ML séparé)")
    df_ml=load_ml()
    if df_ml is not None:
        try: st.dataframe(df_ml.style.highlight_max(subset=['AUC-ROC'],color='#D5F5E3'),width='stretch')
        except: st.dataframe(df_ml,width='stretch')
        st.caption("Métriques sur test set stratifié (20%).")
    else:
        st.info("ℹ️ resultats_modeles_ML.csv introuvable.")
    for fname,cap in [
        ("images/courbes_ROC_comparaison.png","Courbes ROC"),
        ("images/importance_variables_RF_XGB.png","Importance des variables (RF+XGB)"),
        ("images/validation_croisee_boxplot.png","Validation croisée 5-Fold"),
        ("images/matrices_confusion_tous_modeles.png","Matrices de confusion"),
    ]:
        if os.path.exists(fname):
            st.markdown(f"#### {cap}")
            st.image(Image.open(fname),use_container_width=True)

# ── Onglet 6 : Mini-rapport ───────────────────────────────────────────
with tab6:
    st.markdown("### 📄 Mini-rapport")
    if MA:
        rpt=st.session_state.get('rapport','')
        if rpt: st.text(rpt)
        else: st.warning("⚠️ Rapport non disponible.")
    else:
        files=sorted(glob.glob("exports/mini_rapport_PFE_*.txt"))
        if files:
            with open(files[-1],"r",encoding="utf-8") as f: st.text(f.read())
        else: st.warning("⚠️ mini_rapport_PFE_*.txt introuvable.")

# ── Onglet 7 : Analyses approfondies ─────────────────────────────────
with tab7:
    st.markdown("### 🔬 Analyses approfondies")

    # Heatmap
    st.markdown("#### 🔥 Heatmap – Corrélations entre facteurs environnementaux")
    st.caption("Facteurs variant ensemble (corrélations ≥ 0.3 affichées).")
    if MA:
        fb=st.session_state.get('figs_bytes',{}).get('heatmap')
        if fb: st.image(fb,use_container_width=True)
        else:  st.info("ℹ️ Heatmap non disponible.")
    else:
        if os.path.exists("images/heatmap_facteurs_env.png"):
            st.image(Image.open("images/heatmap_facteurs_env.png"),use_container_width=True)
        else: st.warning("⚠️ heatmap_facteurs_env.png introuvable.")

    st.divider()

    # Stratification
    st.markdown("#### 🗺️ Stratification par ville – OR ajustés")
    st.caption("Compare l'effet de chaque facteur entre Casablanca et Jadida.")
    if MA:
        dft7 = st.session_state.get('tableaux', {}).get('T7')
    else:
        dft7 = load_sheet("T7_Stratification_ville")
    if dft7 is not None and not dft7.empty:
        st.dataframe(dft7, width='stretch')
    else:
        st.info("ℹ️ Tableau de stratification (T7) : aucun facteur significatif (p<0,05) pour stratification par ville.")

    st.divider()

    # SHAP
    st.markdown("#### 🧠 Interprétabilité SHAP")
    st.caption("SHAP quantifie la contribution de chaque variable à chaque prédiction individuelle.")
    if MA:
        fb=st.session_state.get('figs_bytes',{}).get('shap')
        if fb: st.image(fb,use_container_width=True)
        else:  st.info("ℹ️ SHAP non disponible (vérifiez que shap est installé : pip install shap).")
    else:
        shap_files = ['images/shap_summary.png', 'images/shap_importance.png', 'images/shap_beeswarm.png']
        found_any = False
        for fname in shap_files:
            if os.path.exists(fname):
                found_any = True
                st.markdown(f"**SHAP – {fname.replace('.png','').replace('_',' ').title()}**")
                st.image(Image.open(fname),use_container_width=True)
        if not found_any:
            st.info("ℹ️ Aucun fichier SHAP trouvé. Exécutez le pipeline pour générer shap_summary.png.")

# ======================================================================
# 8. PIED DE PAGE
# ======================================================================
st.sidebar.divider()
st.sidebar.caption("📌 PFE Data Science – Institut Pasteur du Maroc")
st.sidebar.caption(f"🚀 {pd.Timestamp.now().strftime('%d/%m/%Y')}")
